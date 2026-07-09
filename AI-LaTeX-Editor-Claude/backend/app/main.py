import io
import tempfile
import zipfile
from pathlib import Path
from urllib.parse import quote

from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

from app.ai_providers import get_provider
from app.compile_service import trigger_compile
from app.latex import (
    LATEX_SYSTEM_PROMPT,
    decode_text_bytes,
    extract_docx_text,
    extract_fenced_files,
    extract_first_code_block,
    generate_prompt,
    normalize_main_tex,
    project_context,
    repair_mojibake,
    read_template_zip,
    sse,
    table_from_csv_bytes,
    table_rows_to_latex,
)
from app.material_templates import seed_material_templates
from app.schemas import (
    AIChatRequest,
    CompileRequest,
    CompileResponse,
    FileCreate,
    FileResponse as FileSchema,
    FileUpdate,
    ProjectCreate,
    ProjectResponse,
    ProjectUpdate,
    TemplateDetailResponse,
    TemplateResponse,
    TokenResponse,
    UserLogin,
    UserRegister,
    UserResponse,
)
from app.security import create_token, decode_token, hash_password, verify_password
from app.settings import get_settings
from app.store import detect_file_type, normalize_path, store

settings = get_settings()
for folder in (settings.data_dir, settings.upload_dir, settings.compile_dir, settings.template_dir):
    Path(folder).mkdir(parents=True, exist_ok=True)

LATEX_SOURCE_EXTENSIONS = {
    ".tex",
    ".bib",
    ".sty",
    ".cls",
    ".bst",
    ".bbx",
    ".cbx",
    ".cfg",
    ".def",
    ".clo",
    ".dtx",
    ".ins",
    ".ltx",
    ".lua",
}
LATEX_SOURCE_NAMES = {"Makefile", "latexmkrc", ".latexmkrc"}

app = FastAPI(title=settings.app_name, version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
app.mount("/static/uploads", StaticFiles(directory=settings.upload_dir), name="uploads")


@app.on_event("startup")
async def startup_seed_templates():
    if settings.seed_material_templates:
        seed_material_templates()


def user_response(user: dict) -> UserResponse:
    return UserResponse(
        id=user["id"],
        username=user.get("username"),
        email=user.get("email"),
        is_admin=bool(user.get("is_admin")),
        is_guest=bool(user.get("is_guest")),
    )


def auth_user_id(authorization: str | None) -> str | None:
    if authorization and authorization.startswith("Bearer "):
        return decode_token(authorization[7:])
    return None


def require_user_id(authorization: str | None = Header(default=None)) -> str:
    user_id = auth_user_id(authorization)
    if not user_id or not store.get_user(user_id):
        raise HTTPException(status_code=401, detail="Not authenticated")
    return user_id


def optional_user_id(authorization: str | None = Header(default=None)) -> str | None:
    user_id = auth_user_id(authorization)
    return user_id if user_id and store.get_user(user_id) else None


def compile_payload(job: dict) -> CompileResponse:
    pdf_url = f"/api/projects/{job['project_id']}/pdf" if job.get("status") == "done" else None
    return CompileResponse(
        id=job["id"],
        status=job["status"],
        log=job.get("log"),
        pdf_url=pdf_url,
        error_msg=job.get("error_msg"),
        created_at=job.get("created_at"),
    )


def llm_from_request(request: Request):
    api_key = request.headers.get("X-LLM-Key")
    provider_name = request.headers.get("X-LLM-Provider", "openai")
    if not api_key:
        raise HTTPException(status_code=400, detail="X-LLM-Key header is required for AI operations")
    return get_provider(api_key, provider_name)


def upload_destination(project_id: str, path: str) -> Path:
    safe_path = normalize_path(path)
    root = Path(settings.upload_dir) / project_id
    destination = root / safe_path
    destination.parent.mkdir(parents=True, exist_ok=True)
    return destination


def asset_url(project_id: str, path: str) -> str:
    return "/static/uploads/" + quote(project_id) + "/" + "/".join(quote(part) for part in normalize_path(path).split("/"))


def upload_source(project_id: str, path: str) -> Path:
    return Path(settings.upload_dir) / project_id / normalize_path(path)


def is_latex_source(path: str) -> bool:
    item = Path(path)
    return item.suffix.lower() in LATEX_SOURCE_EXTENSIONS or item.name in LATEX_SOURCE_NAMES


async def ensure_project_pdf(project_id: str) -> Path:
    current = store.latest_pdf(project_id)
    if current and Path(current).exists():
        return Path(current)

    job = await trigger_compile(project_id)
    pdf_path = job.get("pdf_path")
    if job.get("status") == "done" and pdf_path and Path(pdf_path).exists():
        return Path(pdf_path)

    raise HTTPException(status_code=409, detail=job.get("error_msg") or "No compiled PDF available")


@app.get("/api/health")
async def health():
    return {"status": "ok", "storage": "local-json"}


@app.post("/api/auth/register", response_model=TokenResponse)
async def register(data: UserRegister):
    try:
        user = store.create_user(data.username, hash_password(data.password), data.email, is_guest=False)
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc))
    return TokenResponse(access_token=create_token(user["id"]), user=user_response(user))


@app.post("/api/auth/login", response_model=TokenResponse)
async def login(data: UserLogin):
    user = store.get_user_by_username(data.username)
    if not user or not verify_password(data.password, user.get("password_hash")):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    return TokenResponse(access_token=create_token(user["id"]), user=user_response(user))


@app.post("/api/auth/guest", response_model=TokenResponse)
async def guest_login():
    user = store.create_user(None, None, is_guest=True)
    return TokenResponse(access_token=create_token(user["id"]), user=user_response(user))


@app.get("/api/auth/me", response_model=UserResponse)
async def me(user_id: str = Depends(require_user_id)):
    user = store.get_user(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user_response(user)


@app.get("/api/projects", response_model=list[ProjectResponse])
async def list_projects(user_id: str | None = Depends(optional_user_id)):
    return store.list_projects(user_id)


@app.post("/api/projects", response_model=ProjectResponse, status_code=201)
async def create_project(data: ProjectCreate, user_id: str | None = Depends(optional_user_id)):
    if data.template_id and not store.get_template(data.template_id):
        raise HTTPException(status_code=404, detail="Template not found")
    return store.create_project(data.name, user_id=user_id, template_id=data.template_id)


@app.get("/api/projects/{project_id}", response_model=ProjectResponse)
async def get_project(project_id: str):
    project = store.get_project(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project


@app.put("/api/projects/{project_id}", response_model=ProjectResponse)
async def update_project(project_id: str, data: ProjectUpdate):
    project = store.update_project(project_id, data.model_dump(exclude_none=True))
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project


@app.delete("/api/projects/{project_id}", status_code=204)
async def delete_project(project_id: str):
    if not store.delete_project(project_id):
        raise HTTPException(status_code=404, detail="Project not found")


@app.get("/api/projects/{project_id}/files", response_model=list[FileSchema])
async def list_files(project_id: str):
    if not store.get_project(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return store.list_files(project_id)


@app.post("/api/projects/{project_id}/files/upload", response_model=FileSchema, status_code=201)
async def upload_file(project_id: str, file: UploadFile = File(...)):
    if not store.get_project(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    raw = await file.read()
    filename = normalize_path(file.filename or "untitled")
    path = store.unique_path(project_id, filename)
    file_type = detect_file_type(path)
    ext = Path(path).suffix.lower()
    if file_type == "img" or ext in {".pdf", ".docx", ".xlsx", ".zip"}:
        dest = upload_destination(project_id, path)
        dest.write_bytes(raw)
        content = None
        size = len(raw)
        created = store.create_file(project_id, path, content, file_type=file_type, size=size)
        if ext == ".docx":
            try:
                extracted = extract_docx_text(raw)
            except Exception:
                extracted = ""
            if extracted.strip():
                text_path = store.unique_path(project_id, f"sources/{Path(path).stem}.txt")
                store.create_file(project_id, text_path, extracted, file_type="other", size=len(extracted))
        return created
    else:
        content = decode_text_bytes(raw)
        size = len(content)
    return store.create_file(project_id, path, content, file_type=file_type, size=size)


@app.post("/api/projects/{project_id}/files", response_model=FileSchema, status_code=201)
async def create_file(project_id: str, data: FileCreate):
    if not store.get_project(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    try:
        return store.create_file(project_id, data.path, data.content, data.file_type)
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc))


@app.get("/api/projects/{project_id}/files/{file_id}", response_model=FileSchema)
async def get_file(project_id: str, file_id: str):
    item = store.get_file(file_id)
    if not item or item["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="File not found")
    return item


@app.get("/api/projects/{project_id}/files/{file_id}/preview")
async def preview_file(project_id: str, file_id: str):
    item = store.get_file(file_id)
    if not item or item["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="File not found")

    path = item["path"]
    ext = Path(path).suffix.lower()
    source = upload_source(project_id, path)
    base = {
        "id": item["id"],
        "path": path,
        "file_type": item["file_type"],
        "size": item["size"],
    }

    if item.get("content") is not None:
        return {**base, "preview_type": "text", "text": repair_mojibake(item.get("content") or "")}

    if ext in {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg"}:
        return {**base, "preview_type": "image", "asset_url": asset_url(project_id, path)}

    if ext == ".pdf":
        return {**base, "preview_type": "pdf", "asset_url": asset_url(project_id, path)}

    if ext == ".docx" and source.exists():
        try:
            text = extract_docx_text(source.read_bytes())
        except Exception as exc:
            text = f"Unable to extract DOCX preview: {exc}"
        return {**base, "preview_type": "text", "text": repair_mojibake(text)}

    if ext == ".xlsx" and source.exists():
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(source, data_only=True, read_only=True)
            sheet = workbook.active
            rows = []
            for row in sheet.iter_rows(max_row=60, values_only=True):
                rows.append([cell if cell is not None else "" for cell in row[:20]])
            return {**base, "preview_type": "table", "rows": rows}
        except Exception as exc:
            return {**base, "preview_type": "text", "text": f"Unable to extract spreadsheet preview: {exc}"}

    if source.exists() and ext in {".txt", ".md", ".csv", ".log", ".json"}:
        return {**base, "preview_type": "text", "text": repair_mojibake(decode_text_bytes(source.read_bytes()))}

    return {**base, "preview_type": "download", "asset_url": asset_url(project_id, path)}


@app.put("/api/projects/{project_id}/files/{file_id}", response_model=FileSchema)
async def update_file(project_id: str, file_id: str, data: FileUpdate):
    item = store.get_file(file_id)
    if not item or item["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="File not found")
    updated = store.update_file(file_id, data.content)
    return updated


@app.delete("/api/projects/{project_id}/files/{file_id}", status_code=204)
async def delete_file(project_id: str, file_id: str):
    item = store.get_file(file_id)
    if not item or item["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="File not found")
    store.delete_file(file_id)


@app.post("/api/projects/{project_id}/compile", response_model=CompileResponse)
async def compile_project(project_id: str, _: CompileRequest):
    if not store.get_project(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    return compile_payload(await trigger_compile(project_id))


@app.get("/api/projects/{project_id}/compile/{job_id}", response_model=CompileResponse)
async def compile_status(project_id: str, job_id: str):
    job = store.get_compile_job(job_id)
    if not job or job["project_id"] != project_id:
        raise HTTPException(status_code=404, detail="Compile job not found")
    return compile_payload(job)


@app.get("/api/projects/{project_id}/pdf")
async def project_pdf(project_id: str):
    path = store.latest_pdf(project_id)
    if not path or not Path(path).exists():
        raise HTTPException(status_code=404, detail="No compiled PDF available")
    return FileResponse(path, media_type="application/pdf", filename="main.pdf")


@app.post("/api/projects/{project_id}/ai/chat")
async def ai_chat(project_id: str, data: AIChatRequest, request: Request):
    if not store.get_project(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    provider = llm_from_request(request)
    files = store.list_files(project_id)
    context, _, _ = project_context(files)
    active_hint = data.target_file or (files[0]["path"] if files else "main.tex")
    system = f"{LATEX_SYSTEM_PROMPT}\n\nProject files:\n{context}\n\nCurrent file: {active_hint}"
    messages = [{"role": "system", "content": system}, *data.history[-10:], {"role": "user", "content": data.message}]

    async def event_stream():
        yield sse({"type": "start"})
        try:
            async for token in provider.chat_stream(messages, data.model):
                yield sse({"type": "token", "content": token})
            yield sse({"type": "done"})
        except Exception as exc:
            yield sse({"type": "error", "message": str(exc)})

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.post("/api/projects/{project_id}/ai/generate")
async def ai_generate(project_id: str, request: Request):
    if not store.get_project(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    provider = llm_from_request(request)
    model = request.headers.get("X-LLM-Model", "gpt-4o")
    prompt = generate_prompt(store.list_files(project_id))
    messages = [{"role": "system", "content": LATEX_SYSTEM_PROMPT}, {"role": "user", "content": prompt}]

    async def event_stream():
        yield sse({"type": "start"})
        full: list[str] = []
        try:
            async for token in provider.chat_stream(messages, model):
                full.append(token)
                yield sse({"type": "token", "content": token})
            response = "".join(full)
            generated_files = extract_fenced_files(response)
            if not generated_files:
                generated_files = [{"path": "main.tex", "content": normalize_main_tex(response)}]
            if not any(item["path"] == "main.tex" for item in generated_files):
                generated_files.insert(0, {"path": "main.tex", "content": normalize_main_tex(response)})
            for item in generated_files:
                content = repair_mojibake(item["content"])
                if item["path"] == "main.tex":
                    content = normalize_main_tex(content)
                existing = store.get_file_by_path(project_id, item["path"])
                if existing:
                    store.update_file(existing["id"], content)
                else:
                    store.create_file(project_id, item["path"], content, detect_file_type(item["path"]))
                yield sse({"type": "saved", "path": item["path"]})
            yield sse({"type": "done"})
        except Exception as exc:
            yield sse({"type": "error", "message": str(exc)})

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.post("/api/projects/{project_id}/ai/import")
async def ai_import(project_id: str, request: Request, file: UploadFile = File(...)):
    if not store.get_project(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    raw = await file.read()
    filename = normalize_path(file.filename or "untitled")
    ext = Path(filename).suffix.lower()
    stem = Path(filename).stem
    created: list[dict] = []

    if ext in {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg"}:
        safe_image_path = store.unique_path(project_id, filename)
        upload_destination(project_id, safe_image_path).write_bytes(raw)
        created.append(store.create_file(project_id, safe_image_path, None, file_type="img", size=len(raw)))
        figure = "\n".join(
            [
                r"\begin{figure}[htbp]",
                r"\centering",
                rf"\includegraphics[width=0.8\textwidth]{{{safe_image_path}}}",
                r"\caption{}",
                rf"\label{{fig:{stem}}}",
                r"\end{figure}",
            ]
        )
        created.append(store.create_file(project_id, store.unique_path(project_id, f"fig_{stem}.tex"), figure, "tex"))
        return {"files": created}

    if ext == ".csv":
        latex = table_from_csv_bytes(raw, filename)
        return store.create_file(project_id, store.unique_path(project_id, f"table_{stem}.tex"), latex, "tex")

    if ext == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw), data_only=True)
        sheet = workbook.active
        rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        latex = table_rows_to_latex(rows, stem)
        return store.create_file(project_id, store.unique_path(project_id, f"table_{stem}.tex"), latex, "tex")

    if ext == ".docx":
        import subprocess

        with tempfile.NamedTemporaryFile(suffix=".docx", delete=False) as tmp:
            tmp.write(raw)
            tmp_path = tmp.name
        try:
            result = subprocess.run(["pandoc", tmp_path, "--to", "latex", "--wrap=none"], capture_output=True, text=True, timeout=30)
            if result.returncode != 0:
                raise HTTPException(status_code=500, detail="pandoc failed. Install pandoc or upload the DOCX as a staged asset.")
            latex = result.stdout
            return store.create_file(project_id, store.unique_path(project_id, f"{stem}.tex"), latex, "tex")
        finally:
            Path(tmp_path).unlink(missing_ok=True)

    provider = llm_from_request(request)
    model = request.headers.get("X-LLM-Model", "gpt-4o")
    text = decode_text_bytes(raw)
    messages = [
        {"role": "system", "content": LATEX_SYSTEM_PROMPT},
        {"role": "user", "content": f"Convert this source material to a complete LaTeX file:\n\n{text}"},
    ]
    full: list[str] = []
    async for token in provider.chat_stream(messages, model):
        full.append(token)
    content = extract_first_code_block("".join(full))
    return store.create_file(project_id, store.unique_path(project_id, f"{stem}.tex"), content, "tex")


@app.get("/api/templates", response_model=list[TemplateResponse])
async def list_templates():
    return store.list_templates()


@app.get("/api/templates/{template_id}", response_model=TemplateDetailResponse)
async def get_template(template_id: str):
    template = store.get_template(template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    return template


@app.post("/api/admin/templates", response_model=TemplateResponse, status_code=201)
async def create_template(
    file: UploadFile = File(...),
    name: str = Form(...),
    description: str = Form(""),
    category: str = Form("other"),
    language: str = Form("universal"),
):
    zip_bytes = await file.read()
    try:
        files = read_template_zip(zip_bytes)
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="Template upload must be a zip file")
    if not files:
        raise HTTPException(status_code=400, detail="No text LaTeX files found in template zip")
    return store.create_template(
        {"name": name, "description": description, "category": category, "language": language},
        files,
    )


@app.delete("/api/admin/templates/{template_id}", status_code=204)
async def hide_template(template_id: str):
    template = store.get_template(template_id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    data = store._load()
    for item in data["templates"]:
        if item["id"] == template_id:
            item["is_public"] = False
    store._save(data)


@app.get("/api/projects/{project_id}/export/tex")
async def export_tex(project_id: str):
    if not store.get_project(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    pdf_path = await ensure_project_pdf(project_id)
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in store.list_files(project_id):
            if item.get("content") is not None and is_latex_source(item["path"]):
                zf.writestr(item["path"], item.get("content") or "")
        zf.write(pdf_path, "main.pdf")
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/zip",
        headers={"Content-Disposition": 'attachment; filename="latex-source-and-pdf.zip"'},
    )


@app.get("/api/projects/{project_id}/export/pdf")
async def export_pdf(project_id: str):
    if not store.get_project(project_id):
        raise HTTPException(status_code=404, detail="Project not found")
    path = await ensure_project_pdf(project_id)
    return FileResponse(path, media_type="application/pdf", filename="main.pdf")
